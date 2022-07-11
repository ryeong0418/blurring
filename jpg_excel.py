import os
import glob
import openpyxl as xl

main_path="C:/Users/USER/Desktop/6월1주차"
excel_path="C:/Users/USER/Desktop/6월1주차/정제데이터.xlsx"

#파일명 기입
wb=xl.load_workbook(excel_path,data_only=True)
sheet=wb["7-2 정제시트(수집)"]

#파일명 기입한 거 읽어오기
ws=wb.active
head_row=ws['3']

path_list=[]
total_list=[]
blur_list=[]


#파일명, 확장자 기입하기

def jpg_excel(j):

    noise=["FF","RF","RL","UD","LE","DU","SC","FI","LB","RB"]
    gt_list=[]
    noise_list=[]
    syn_list=[]


    for path,dir,files in os.walk(j):
        for file in files:
            if file.endswith(".jpg"):
                name=file.split(".")[0]
                noise_syn=name.split("_")[4]

                if len(file)==27 and noise_syn=="GT":
                    gt_list.append(name)       

                if len(file)==27 and noise_syn in noise:
                    noise_list.append(name)             

                if len(file)==30:
                    syn_list.append(name)
            
    gt_list.extend(noise_list)
    gt_list.extend(syn_list)
 
    return gt_list      

for path,dir,files in os.walk(main_path):
    path_list.append(path)

for i in range(1,len(path_list)):
    filename=path_list[i].split("\\")[1]   
    for i in jpg_excel(path_list[i]):
        total_list.append(i)

for seq,name in enumerate(total_list):
    sheet.cell(row=4+seq,column=2,value=name) # 파일명 기입
    sheet.cell(row=4+seq,column=6,value="jpg") # 확장자 기입
    print(sheet.cell(row=4+seq,column=2,value=name))


    wb.save(excel_path)

# 블러 기입하기


for i in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    if i[1].value is not None:
        noise=["FF","RF","RL","UD","LE","DU","SC","FI","LB","RB"]
        filename=i[1].value
        noise_type=filename.split("_")[4]
              
        if noise_type == "GT":
            i[2].value=noise_type

        if noise_type in noise:
            i[2].value=noise_type
        
        if len(filename)==26:
            syn_type=filename.split("_")[4]+"_"+filename.split("_")[5]
            i[3].value=syn_type

            
            wb.save(excel_path)

print("완료")

