import openpyxl as xl
import os
import json
from collections import OrderedDict
from PIL import Image
from PIL.ExifTags import TAGS
from fractions import Fraction

jpg_path="C:\\Users\\USER\\Desktop\\copy\\ab-after"
excel_path="C:\\Users\\USER\\Desktop\\excelsh\\스크립트 메타데이터시트 - 복사본.xlsx"

wb=xl.load_workbook(excel_path)
sheet=wb["7-3촬영시트(수집취합 정제전)"]
lowlight_data=OrderedDict()

#파일명 변경, 엑셀에 pair해서 insert

def excelinsert():

    beforename_list=[]
    aftername_list=[]

    for path,dir,files in os.walk(jpg_path):
        i=1
        for file in files:
            if file.endswith(".jpg"):
                filename="D3"+"_"+path.split("\\")[-5]+"_"+path.split("\\")[-4]+path.split("\\")[-3]+"_"+path.split("\\")[-2]+"_"+path.split("\\")[-1]          
                lowlight_data["변경전 파일명"]=file #변경전파일명 SY_220525_0001.jpg
                os.rename(path+"/"+file,path+"/"+filename+str(i).zfill(4)+".jpg")
                lowlight_data["변경후 파일명"]=filename+str(i).zfill(4)+".jpg" #변경후파일명 D3_220511_I01_L06A_E0001.jpg
                i+=1

                for key, value in lowlight_data.items():
                    if key=="변경전 파일명":
                        beforename_list.append(value)                    
                    if key=="변경후 파일명":
                        aftername_list.append(value)


    for seq,name in enumerate(beforename_list):
        sheet.cell(row=4+seq,column=2,value=name)

    for seq,name in enumerate(aftername_list):
        sheet.cell(row=4+seq,column=3,value=name)


    wb.save(excel_path)

excelinsert()

#meta값 엑셀에 insert

def jpgmeta_insert():

    filelist=[]
    datelist=[]
    exposuretime_list=[]
    fnumber_list=[]
    location_list=[]
    place_list=[]
    angle_list=[]
    lowenvironment_list=[]
    bracketing_list=[]

    for path,dir,files in os.walk(jpg_path):
        for file in files:
            if file.endswith(".jpg"):
                filename=file.split(".")[0] # D3_220511_I01_L06A_E0001.jpg
                filesplit=filename.split("_") #['D3', '220516', 'I01', 'L06A', 'E0001']
                filedate=file.split(".")[0].split("_")[1] #'220516'
                final_filedate="20"+filedate[:2]+"-"+filedate[2:4]+"-"+filedate[4:6] #2022-05-16
                datelist.append(str(final_filedate))
                filelist.append(str(filename))
                location_list.append(str(filesplit[2][0])) #I,O
                place_list.append(str(filesplit[2][1:3])) #05 
                angle_list.append(str(filesplit[4][0]))

                lowenvironment_list.append(str(filesplit[3][0:3]))

                bracketing_list.append(str(filesplit[3][-1]))

    #조리개수치, 노출시간
    for path,dir,files in os.walk(jpg_path):
        for file in files:
            if file.endswith(".jpg"):
                a=Image.open(path+"\\"+file)
                info=a._getexif()
                for tag,value in info.items(): 
                    decoded=TAGS.get(tag,tag)
                    if decoded=="ExposureTime":
                        exposuretime_list.append(str(Fraction(value)))                
                    if decoded=="FNumber":
                        fnumber_list.append("F"+"/"+str(value))


    #excel에 insert
    # for seq,name in enumerate(filelist):
    #     sheet.cell(row=4+seq,column=15,value=name)

    for seq,name in enumerate(datelist):
        sheet.cell(row=4+seq,column=7,value=name)

    for seq,name in enumerate(exposuretime_list):
        sheet.cell(row=4+seq,column=10,value=name)

    for seq,name in enumerate(fnumber_list):
        sheet.cell(row=4+seq,column=9,value=name)

    for seq,name in enumerate(location_list):
        sheet.cell(row=4+seq,column=5,value=name)

    for seq,name in enumerate(place_list):
        sheet.cell(row=4+seq,column=6,value=name)

    for seq,name in enumerate(angle_list):
        sheet.cell(row=4+seq,column=13,value=name)

    for seq,name in enumerate(lowenvironment_list):
        sheet.cell(row=4+seq,column=11,value=name)

    for seq,name in enumerate(bracketing_list):
        sheet.cell(row=4+seq,column=12,value=name)
    
    
    wb.save(excel_path)

jpgmeta_insert()


