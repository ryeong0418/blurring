import os
import json
import openpyxl as xl

excel_path='D:/blurring-linkflow/6m1w/7-2_RawData_촬영일정_6월 1주차_220602.xlsx'
source_path='D:/blurring-linkflow/6m1w/6월_1주차_정제리스트_수량_220602.xlsx'
json_save_path='D:/blurring-linkflow/6m1w/6월 1주차(RS_JSON)'


source_wb=xl.load_workbook(source_path,data_only=True)
source_ws=source_wb.active
head_row=source_ws['3'] #행 뽑는거 

wb=xl.load_workbook(excel_path,data_only=True)
ws=wb.active
head_row=ws['3'] #행 뽑는거 

raw_list=[]
source_list=[]

for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    raw_data_id=str(row[1].value)
    copyrighter="㈜미디어"
    location=str(row[3].value)
    lux_level=str(row[4].value)
    objectid=str(row[5].value)
    date=str(row[6].value)
    length=str(row[7].value)
    resolution=str(row[8].value)
    fps=str(row[9].value)
    fstop=str(row[10].value)
    exposuretime=str(row[11].value)
    angle=str(row[12].value)
    fileextension=str(row[13].value)

    raw_dict={

        "Raw_Data_Info.":{
            "Raw_Data_ID":raw_data_id,
            "Copyrighter":copyrighter,
            "Location":location,
            "Lux_Level":lux_level,
            "Object_ID":objectid,
            "Date":date[:10],
            "Length":length,
            "Resolution":resolution,
            "FPS":fps,
            "F-Stop":fstop,
            "Exposure_Time":exposuretime,
            "Angle":angle,
            "File_Extension":fileextension,
            }
        }

    #print(raw_dict)
    raw_list.append(raw_dict)


for row in source_ws.iter_rows(min_row=4, max_row=source_ws.max_row, min_col=1, max_col=source_ws.max_column):
    source_data_id=str(row[1].value)
    noise_type=str(row[2].value)
    synthesis_type=str(row[3].value)
    center_coordinate=str(row[4].value)
    fileextension=str(row[5].value)

    os.makedirs(json_save_path+"/"+source_data_id[:20],exist_ok=True)

    source_dict={

        "Source_Data_Info.":{            
            "Source_Data_ID":source_data_id,
            "Noise_type":noise_type,
            "Synthesis_type":synthesis_type,
            "Center_coordinate":center_coordinate,
            "File_Extension":fileextension
        }    
    }

    source_list.append(source_dict)
    
for raw in raw_list:
    for source in source_list:
        source_id=source['Source_Data_Info.']['Source_Data_ID']
        if raw['Raw_Data_Info.']['Raw_Data_ID']==source_id.split("_")[0]+"_"+source_id.split("_")[1]+"_"+source_id.split("_")[2]+"_"+source_id.split("_")[3]:
            raw.update(source)

               
            with open(os.path.join(json_save_path+"/"+source_id[:20],source_id+".json"),"w",encoding="utf-8") as json_file:
                json.dump(raw,json_file,indent=4,ensure_ascii=False,default=str)



