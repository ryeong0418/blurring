import os
import openpyxl as xl

excel_path="D:/occ/6월 4주차_220623/7-1.xlsx"
data_path=""

wb=xl.load_workbook(excel_path,data_only=True)
ws=wb["7-1 촬영시트(수집)"]
head_row=ws['3']

#print(head_row)

filename_list=[]
copyrighter_list=[]
location_list=[]
target_list=[]
place_list=[]
date_list=[]
length_list=[]
resolution_list=[]
fps_list=[]
fstop_list=[]
extime_list=[]
angle_list=[]
file_extension_list=[]

place_total=["('I','01','집')","('I','02','스튜디오')","('I','03','식당')","('I','04','사무실')","('I','05','카페')","('O','01','운동장')","('O','02','산책로')",
    "('O','03','도로')","('O','04','주차장')","('O','05','광장')","('O','06','선착장')","('O','07','공원')"]

for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    #print(row)
    if row[1].value is not None:
        filename_list.append(str(row[1].value))
        copyrighter_list.append(str(row[2].value))
        location_list.append(str(row[3].value))
        target_list.append(str(row[4].value))
        place_list.append(str(row[5].value))
        date_list.append(str(row[6].value))
        length_list.append(str(row[7].value))
        resolution_list.append(str(row[8].value))
        fps_list.append(str(row[9].value))
        fstop_list.append(str(row[10].value))
        extime_list.append(str(row[11].value))
        angle_list.append(str(row[12].value))
        file_extension_list.append(str(row[13].value))



# 파일명과 실내 및 실외 비교
for i in range(len(filename_list)):
    if filename_list[i].split("_")[2][0] != location_list[i]:
        print("파일명과 실내,실외 오류: "+str("No.")+str(i+1))


#copyrighter 오류
for i in range(len(copyrighter_list)):
    if copyrighter_list[i] != "㈜미디어":
        print("Copyrighter오류: "+str("No.")+str(i+1))


#파일명과 폐색유형 비교
for i in range(len(filename_list)):
    if filename_list[i].split("_")[2][1:3] != target_list[i]:
        print("파일명과 폐색유형 오류: "+str("No.")+str(i+1))


for i in range(len(filename_list)):
    i_o=filename_list[i].split("_")[2][0]
    number=filename_list[i].split("_")[2][3:5]
    place_list[i]
    mk_place="("+"'"+i_o+"'"+","+"'"+number+"'"+","+"'"+place_list[i]+"'"+")"

    if mk_place not in place_total:
        print("파일명과 촬영장소 id 오류: "+str("No.")+str(i+1))

#파일명과 촬영일자 비교
for i in range(len(filename_list)):
    d=filename_list[i].split("_")[1]
    date="20"+d[0:2]+"-"+d[2:4]+"-"+d[4:6]
    if date != date_list[i][:10]:
        print("파일명과 날짜 오류: "+str("No.")+str(i+1))


#길이 비교
for i in range(len(length_list)):
    if len(length_list[i])!=1 and len(length_list[i])!=2:
        print("length 오류 : "+str("No.")+str(i+1))


# 해상도 오류
for i in range(len(resolution_list)):
    if resolution_list[i] != "1920, 1080":
        print("Copyrighter오류: "+str("No.")+str(i+1))


#fps 오류
for i in range(len(filename_list)):
    if filename_list[i].split("_")[2][2] == "S" and fps_list[i] != "5":
        print("fps오류: "+str("No.")+str(i+1))

    if filename_list[i].split("_")[2][2] == "A" and fps_list[i] != "30":
        print("fps오류: "+str("No.")+str(i+1))


#f-stop 오류
for i in range(len(fstop_list)):
    if fstop_list[i][0]=="F" and 1.8<=float(fstop_list[i][1:])<=12:
        pass
    else:
        print("fstop오류: "+str("No.")+str(i+1))


#EXPOSURE_TIME 오류
for i in range(len(extime_list)):
    if extime_list[i][0]=="1" and extime_list[i][1]=="/" and 40<=int(extime_list[i][2:])<=1043:
        pass
    else:
        print("exposure_time오류: "+str("No.")+str(i+1))


#angle 오류
for i in range(len(angle_list)):
    if angle_list[i] not in ["Eye","High","Low"]:
        print("angle오류: "+str("No.")+str(i+1))

    if filename_list[i].split("_")[3][0] != angle_list[i][0]:
        print("angle오류: "+str("No.")+str(i+1))


# 확장자 오류
for i in range(len(file_extension_list)):
    if file_extension_list[i] != "mp4":
        print("확장자 오류: "+str("No.")+str(i+1))
