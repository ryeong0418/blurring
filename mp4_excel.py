import os
import openpyxl as xl

data_path="C:/Users/USER/Desktop/mp4_test/6월 1주차(원시데이터)"
excel_path="C:/Users/USER/Desktop/mp4_test/7-2_링크플로우_RawData_촬영일정_6월 1주차_220602.xlsx"

wb=xl.load_workbook(excel_path,data_only=True)
sheet=wb["7-2 촬영시트(수집)"]

ws=wb.active
head_row=ws['3']

object_dict={'01':'화분','02':'시계','03':'스마트폰','04':'공기청정기','05':'메뉴판','06':'책표지','07':'사과','08':'칠판','09':'달력','10':'컵',
'11':'리코더','12':'자전거','13':'실외기','14':'휴대용게임기','15':'음료수','16':'전자레인지','17':'안내표지판','18':'자동차','19':'그네','20':'우산',
'21':'카메라','22':'야구배트','23':'미끄럼틀','24':'스피커','25':'튜브','26':'가방','27':'라켓','28':'헬멧','29':'유모차','30':'안전봉',
'31':'안경','32':'모자','33':'테이블','34':'프라이팬','35':'피자','36':'스탠드조명','37':'라바콘','38':'오토바이','39':'TV','40':'거울',
'41':'공','42':'사람','43':'랜턴','44':'액자','45':'RC카','46':'벤치','47':'풍선','48':'휴지통','49':'주전자','50':'소화기',
'51':'야구글러브','52':'시소','53':'근력운동기구','54':'노트북','55':'구두','56':'가위','57':'피망','58':'세탁기','59':'체스판','60':'프린터기',
'61':'소파','62':'냄비','63':'칼','64':'냉장고','65':'피아노','66':'자판기','67':'수박','68':'다리미','69':'와인병','70':'선풍기',
'71':'국자','72':'기타','73':'신문','74':'리모컨','75':'포도','76':'에어컨','77':'토스터기','78':'의자','79':'모니터','80':'침대'}

mp4_error=[]
filename_list=[]
location_list=[]
date_list=[]
angle_list=[]
id_list=[]
object_list=[]

for path,dir,files in os.walk(data_path):
    for file in files:

        if file.endswith(".mp4"):
            filename=file.split(".")[0]
            # 파일명 리스트 생성
            filename_list.append(filename)

            # Location 리스트 생성
            l=filename.split("_")[2][0]
            location_list.append(l)

            # 촬영날짜 리스트 생성
            d=filename.split("_")[1]
            date="20"+d[0:2]+"-"+d[2:4]+"-"+d[4:6]
            date_list.append(date)

            #angle 리스트 생성
            a=filename.split("_")[3][0]
            if a=="E":   
                angle_list.append("Eye")                
            if a=="H":  
                angle_list.append("High")
            if a=="L":
                angle_list.append("Low")

            #object_id 리스트 생성
            object_id=filename.split("_")[2][2:4]
            id_list.append(object_id)
               
        else:
            mp4_error.append(file)

print("확장자 오류:",mp4_error)

# 객체 id 리스트 만들기

for id in id_list:
    if id in object_dict.keys():
        object_list.append(object_dict[id])


for seq,name in enumerate(filename_list):
    sheet.cell(row=4+seq,column=2,value=name) # 파일명 기입
    sheet.cell(row=4+seq,column=3,value="㈜미디어그룹사람과숲") #copyrighter 기입
    sheet.cell(row=4+seq,column=9,value="1920, 1080") # 해상도 기입
    sheet.cell(row=4+seq,column=10,value="30") # FPS 기입
    sheet.cell(row=4+seq,column=14,value="mp4") # 확장자 기입

for seq,name in enumerate(location_list):
    sheet.cell(row=4+seq,column=4,value=name) # 실내, 실외 기입

for seq,name in enumerate(object_list):
    sheet.cell(row=4+seq,column=6,value=name) # 객체id 기입

for seq,name in enumerate(date_list):
    sheet.cell(row=4+seq,column=7,value=name) # 촬영날짜 기입

for seq,name in enumerate(angle_list):
    sheet.cell(row=4+seq,column=13,value=name) # 촬영각도 기입


wb.save(excel_path)
print("완료")


    
