import os
import openpyxl as xl


main_path="D:/blurring_linkflow(merge)/jpg"
excel_path="D:/blurring_linkflow(merge)/jpg/jpg_count.xlsx"

wb=xl.load_workbook(excel_path)
sheet=wb["Sheet1"]

def object_ID():
    
    H_key_list=[]
    H_value_list=[]
    H_filelist=[]
    
    L_key_list=[]
    L_value_list=[]
    L_filelist=[]

    for path,dir,files in os.walk(main_path):
        for filename in files:
            if filename.endswith(".jpg"):
    
                ####### 밝은조도일때 ######
                if filename.split(".")[0].split("_")[2][1]=="H": 
                    H_filelist.append(filename.split("_")[2][2:4]) ##object_id를 H_filelist에 넣기

                
                ####### 어두운 조도 일때 #####
                if filename.split(".")[0].split("_")[2][1]=="L": 
                    L_filelist.append(filename.split("_")[2][2:4]) ##object_id를 L_filelist에 넣기

    h_dict={}  ######밝은 조도 딕셔너리 생성
    l_dict={}  ######어두운 조도 딕셔너리 생성
    h_sum=0
    l_sum=0

    i=0    

    while i < 80:
        i+=1
        h_dict[str(i).zfill(2)]=0 ## 밝은 조도 {'01': 0, '02': 0, '03': 0,,,'80': 0}
        l_dict[str(i).zfill(2)]=0 ## 어두운 조도 {'01': 0, '02': 0, '03': 0,,,'80': 0}

    for i in H_filelist:
        h_dict[i]+=1  ### 밝은 조도 딕셔너리 value값 카운트 하기

    for i in L_filelist:
        l_dict[i]+=1  ### 어두운 조도 딕셔너리 value값 카운트 하기

    for i in h_dict.values():  ### 밝은 조도 총 개수 구하기
        h_sum+=i

    for i in l_dict.values():  ### 어두운 조도 총 개수 구하기
        l_sum+=i

    print("밝은 조도 object_id 카운트 ",h_dict)
    print("밝은 조도 object_id 총개수",h_sum)
    print("="*50)
    print("어두운 조도 object_id 카운트 ",l_dict)
    print("어두운 조도 object_id 총개수",l_sum)
    print("="*50)
    print("객체 총 개수",h_sum+l_sum)


############ 밝은 조도 엑셀에 넣기 ############
    for key, value in h_dict.items():
        H_key_list.append(key)
        H_value_list.append(value)

    for seq,name in enumerate(H_key_list):
        sheet.cell(row=1+seq, column=1,value=name)
    
    for seq,name in enumerate(H_value_list):
        sheet.cell(row=1+seq, column=2,value=name)



############ 어두운 조도 엑셀에 넣기 ############

    for key, value in l_dict.items():
        L_key_list.append(key)
        L_value_list.append(value)

    for seq,name in enumerate(L_key_list):
        sheet.cell(row=1+seq, column=4,value=name)
    
    for seq,name in enumerate(L_value_list):
        sheet.cell(row=1+seq, column=5,value=name)


    wb.save(excel_path)


if __name__=="__main__":
    object_ID()
