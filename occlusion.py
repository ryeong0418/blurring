import os
import json
import openpyxl as xl
import shutil

excel_path=r'D:\일정한폐색2.xlsx'
jpg_path=r'C:\Users\USER\Desktop\D'
new_jpg_path=r'C:\Users\USER\Desktop\D_NEW'

# 1. 엑셀에서 폴더명 추출

wb=xl.load_workbook(excel_path,data_only=True)
ws=wb["Sheet1"]

jpg_list=[]
new_list=[]
jpg_folder_list=[]


for row in ws.iter_rows(min_row=2,max_row=ws.max_row,min_col=2,max_col=2):
    jpg_name=row[0].value 
    #print(jpg_name) # D2_220602_OH15_E0001_AV_01.jpg
    jpg_split=jpg_name.split(".")[0].split("_")
    folder_name=jpg_split[0]+"_"+jpg_split[1]+"_"+jpg_split[2]+"_"+jpg_split[3]
    jpg_list.append(folder_name)

    # 리스트 중복 제거 방법
    for i in jpg_list:
        if i not in new_list:
            new_list.append(i)

print(new_list) #['D1_221009_ISS02_E0008', 'D1_221009_ISS02_E0009']


# 2. 엑셀 폴더명과 일치하는 실제 jpg 폴더명 추출 하여 jpg 이동시키기

for path,dir,files in os.walk(jpg_path):
    jpg_folder_list.append(path) # 'C:\\Users\\USER\\Desktop\\D\\D1_221009_ISS02_E0008',

for i in jpg_folder_list:
    jpg_f=i.split("\\")[-1]
    for j in new_list:
        if jpg_f==j:
            print(i,new_jpg_path+"/"+jpg_f)
            shutil.copytree(i,new_jpg_path+"/"+jpg_f)



